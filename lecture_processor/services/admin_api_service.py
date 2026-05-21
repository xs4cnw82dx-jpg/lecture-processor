"""Business logic handlers for admin APIs."""

from datetime import datetime, timezone

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

from lecture_processor.services import admin_dashboard_service
from lecture_processor.services import admin_support
from lecture_processor.domains.admin import metrics as admin_metrics
from lecture_processor.domains.billing import credits as billing_credits
from lecture_processor.domains.shared import sanitize_excel_cell

ADMIN_CREDIT_GRANT_MAX = 10000
ADMIN_CREDIT_FIELDS_BY_CATEGORY = {
    'lecture': 'lecture_credits_standard',
    'slides': 'slides_credits',
    'interview': 'interview_credits_short',
}
ADMIN_CREDIT_CATEGORY_LABELS = {
    'lecture': 'Lecture Notes',
    'slides': 'Text Extraction',
    'interview': 'Interview Transcription',
}


def _require_admin(app_ctx, request):
    return admin_support.require_admin(app_ctx, request)


def _to_non_negative_float(value, default=0.0):
    return admin_support.to_non_negative_float(value, default=default)


def _to_non_negative_int(value, default=0):
    return admin_support.to_non_negative_int(value, default=default)


def _admin_credit_now(app_ctx):
    return float(app_ctx.time.time())


def _normalize_admin_email(value):
    return billing_credits.normalize_email(value)


def _select_user_doc_for_email(docs, email_normalized):
    for doc in docs:
        data = doc.to_dict() or {}
        if _normalize_admin_email(data.get('email_normalized') or data.get('email')) == email_normalized:
            return doc, data
    if docs:
        doc = docs[0]
        return doc, doc.to_dict() or {}
    return None, {}


def _find_user_doc_by_email(app_ctx, email):
    email_normalized = _normalize_admin_email(email)
    docs = []
    try:
        docs = app_ctx.users_repo.query_by_email_normalized(app_ctx.db, email_normalized, limit=5)
    except Exception as error:
        app_ctx.logger.warning('Admin user email_normalized lookup failed for %s: %s', email_normalized, error)
    doc, data = _select_user_doc_for_email(docs, email_normalized)
    if doc is not None:
        return doc, data

    fallback_candidates = [email, email_normalized]
    seen = set()
    for candidate in fallback_candidates:
        safe_candidate = str(candidate or '').strip()
        if not safe_candidate or safe_candidate in seen:
            continue
        seen.add(safe_candidate)
        try:
            docs = app_ctx.users_repo.query_by_email(app_ctx.db, safe_candidate, limit=5)
        except Exception as error:
            app_ctx.logger.warning('Admin user email lookup failed for %s: %s', safe_candidate, error)
            docs = []
        doc, data = _select_user_doc_for_email(docs, email_normalized)
        if doc is not None:
            return doc, data
    return None, {}


def _serialize_admin_credit_user(app_ctx, uid, user_data):
    user = dict(user_data or {})
    user.setdefault('uid', uid)
    email = str(user.get('email', '') or '').strip()
    email_normalized = _normalize_admin_email(user.get('email_normalized') or email)
    stored_unlimited = billing_credits.normalize_unlimited_credits(user.get('unlimited_credits'))
    effective_unlimited = billing_credits.effective_unlimited_credits(user, runtime=app_ctx)
    raw_credits = {
        'lecture_credits_standard': int(user.get('lecture_credits_standard', 0) or 0),
        'lecture_credits_extended': int(user.get('lecture_credits_extended', 0) or 0),
        'slides_credits': int(user.get('slides_credits', 0) or 0),
        'interview_credits_short': int(user.get('interview_credits_short', 0) or 0),
        'interview_credits_medium': int(user.get('interview_credits_medium', 0) or 0),
        'interview_credits_long': int(user.get('interview_credits_long', 0) or 0),
    }
    return {
        'uid': uid,
        'email': email,
        'email_normalized': email_normalized,
        'credits': {
            'lecture': raw_credits['lecture_credits_standard'] + raw_credits['lecture_credits_extended'],
            'slides': raw_credits['slides_credits'],
            'interview': (
                raw_credits['interview_credits_short']
                + raw_credits['interview_credits_medium']
                + raw_credits['interview_credits_long']
            ),
        },
        'raw_credits': raw_credits,
        'unlimited': stored_unlimited,
        'effective_unlimited': effective_unlimited,
        'is_configured_admin': billing_credits.is_configured_admin_identity(uid=uid, email=email, runtime=app_ctx),
        'account_status': str(user.get('account_status', 'active') or 'active'),
        'total_processed': int(user.get('total_processed', 0) or 0),
        'created_at': user.get('created_at', 0),
        'updated_at': user.get('updated_at', 0),
    }


def _serialize_admin_credit_grant(doc):
    data = doc.to_dict() or {}
    return {
        'id': str(getattr(doc, 'id', '') or data.get('grant_id', '') or ''),
        'action': str(data.get('action', '') or ''),
        'uid': str(data.get('uid', '') or ''),
        'email': str(data.get('email', '') or ''),
        'email_normalized': str(data.get('email_normalized', '') or ''),
        'admin_uid': str(data.get('admin_uid', '') or ''),
        'admin_email': str(data.get('admin_email', '') or ''),
        'credits': data.get('credits', {}) if isinstance(data.get('credits'), dict) else {},
        'credit_categories': data.get('credit_categories', {}) if isinstance(data.get('credit_categories'), dict) else {},
        'unlimited_before': billing_credits.normalize_unlimited_credits(data.get('unlimited_before')),
        'unlimited_after': billing_credits.normalize_unlimited_credits(data.get('unlimited_after')),
        'note': str(data.get('note', '') or ''),
        'price_cents': int(data.get('price_cents', 0) or 0),
        'currency': str(data.get('currency', 'eur') or 'eur'),
        'source': str(data.get('source', 'admin') or 'admin'),
        'created_at': data.get('created_at', 0),
    }


def _serialize_admin_credit_record(grant_id, record):
    return _serialize_admin_credit_grant(
        type('GrantDoc', (), {'id': grant_id, 'to_dict': lambda _self: dict(record or {})})()
    )


def _list_admin_credit_grants(app_ctx, *, email='', uid='', limit=20):
    safe_limit = min(max(_to_non_negative_int(limit, default=20), 1), 100)
    try:
        if email:
            docs = app_ctx.admin_credit_grants_repo.list_by_email_recent(
                app_ctx.db,
                _normalize_admin_email(email),
                limit=safe_limit,
                firestore_module=app_ctx.firestore,
            )
        elif uid:
            docs = app_ctx.admin_credit_grants_repo.list_by_uid_recent(
                app_ctx.db,
                str(uid or '').strip(),
                limit=safe_limit,
                firestore_module=app_ctx.firestore,
            )
        else:
            docs = app_ctx.admin_credit_grants_repo.list_recent(
                app_ctx.db,
                limit=safe_limit,
                firestore_module=app_ctx.firestore,
            )
    except Exception as error:
        app_ctx.logger.warning('Admin credit grant ledger lookup failed: %s', error)
        docs = []
    return [_serialize_admin_credit_grant(doc) for doc in docs]


def _coerce_admin_credit_amounts(payload):
    source = payload.get('credits') if isinstance(payload, dict) else {}
    source = source if isinstance(source, dict) else {}
    category_amounts = {}
    for category in billing_credits.CREDIT_CATEGORIES:
        raw_value = source.get(category, 0)
        try:
            amount = int(raw_value or 0)
        except Exception:
            raise ValueError(f'{ADMIN_CREDIT_CATEGORY_LABELS[category]} credits must be a whole number.')
        if amount < 0:
            raise ValueError(f'{ADMIN_CREDIT_CATEGORY_LABELS[category]} credits cannot be negative.')
        if amount > ADMIN_CREDIT_GRANT_MAX:
            raise ValueError(f'{ADMIN_CREDIT_CATEGORY_LABELS[category]} credits cannot exceed {ADMIN_CREDIT_GRANT_MAX:,} in one grant.')
        if amount > 0:
            category_amounts[category] = amount
    if not category_amounts:
        raise ValueError('Enter at least one positive credit amount.')
    return category_amounts


def _coerce_admin_unlimited_updates(payload):
    source = payload.get('unlimited') if isinstance(payload, dict) else {}
    source = source if isinstance(source, dict) else {}
    updates = {}
    for category in billing_credits.CREDIT_CATEGORIES:
        if category in source:
            updates[category] = bool(source.get(category))
    if not updates:
        raise ValueError('Choose at least one unlimited setting to update.')
    return updates


def _admin_actor(decoded_token):
    token = decoded_token if isinstance(decoded_token, dict) else {}
    return {
        'admin_uid': str(token.get('uid', '') or ''),
        'admin_email': _normalize_admin_email(token.get('email', '')),
    }


def _target_is_deleting(user_data):
    user = user_data if isinstance(user_data, dict) else {}
    return (
        str(user.get('account_status', '') or '').strip().lower() == 'deleting'
        or bool(user.get('delete_requested_at', 0))
        or bool(user.get('delete_started_at', 0))
    )


def _note_from_payload(payload):
    return ' '.join(str((payload if isinstance(payload, dict) else {}).get('note', '') or '').split())[:300]


def _commit_admin_credit_batch(app_ctx, user_ref, user_updates, grant_ref, grant_record):
    try:
        batch = app_ctx.db.batch()
        batch.update(user_ref, user_updates)
        batch.set(grant_ref, grant_record)
        batch.commit()
        return
    except AttributeError as error:
        if '_document_path' not in str(error):
            raise
        # Some local harnesses provide lightweight document refs that cannot
        # be used with Firestore's WriteBatch. Production refs stay atomic.
        user_ref.update(user_updates)
        grant_ref.set(grant_record)


def _coerce_usd_to_eur(payload):
    raw_usd_to_eur = payload.get('usd_to_eur')
    raw_eur_usd = payload.get('eur_usd')
    if raw_usd_to_eur not in (None, ''):
        value = _to_non_negative_float(raw_usd_to_eur, default=0.0)
        if value > 0:
            return value
    if raw_eur_usd not in (None, ''):
        eur_usd = _to_non_negative_float(raw_eur_usd, default=0.0)
        if eur_usd > 0:
            return 1.0 / eur_usd
    return 0.93


def _normalize_analysis_filters(payload, runtime=None):
    _ = runtime
    data = payload if isinstance(payload, dict) else {}
    mode = str(data.get('mode', '') or '').strip()
    status = str(data.get('status', '') or '').strip()
    uid = str(data.get('uid', '') or '').strip()
    email = str(data.get('email', '') or '').strip().lower()
    period = admin_metrics.coerce_analysis_period(data.get('period', data.get('window', 'monthly')))
    selected_ids = data.get('job_ids', data.get('selected_job_ids', []))
    if not isinstance(selected_ids, list):
        selected_ids = []
    selected_ids = [str(item or '').strip() for item in selected_ids if str(item or '').strip()]
    return {
        'period': period,
        'mode': mode,
        'status': status,
        'uid': uid,
        'email': email,
        'selection': str(data.get('selection', 'all') or 'all').strip().lower(),
        'job_ids': selected_ids,
        'single_job_id': str(data.get('job_id', '') or '').strip(),
        'usd_to_eur': _coerce_usd_to_eur(data),
    }


def _job_matches_filters(job, normalized_filters):
    mode = normalized_filters.get('mode', '')
    status = normalized_filters.get('status', '')
    uid = normalized_filters.get('uid', '')
    email = normalized_filters.get('email', '')
    if mode and str(job.get('mode', '') or '').strip() != mode:
        return False
    if status and str(job.get('status', '') or '').strip() != status:
        return False
    if uid and str(job.get('uid', '') or '').strip() != uid:
        return False
    if email and str(job.get('email', '') or '').strip().lower() != email:
        return False
    return True


def _select_jobs(filtered_jobs, normalized_filters):
    selected_ids = set(normalized_filters.get('job_ids', []) or [])
    selection = normalized_filters.get('selection', 'all')
    if selected_ids:
        return [job for job in filtered_jobs if str(job.get('job_id', '') or '') in selected_ids]
    if selection == 'one':
        one_job_id = normalized_filters.get('single_job_id', '')
        if not one_job_id:
            return []
        return [job for job in filtered_jobs if str(job.get('job_id', '') or '') == one_job_id]
    return list(filtered_jobs)


def _build_cost_analysis_payload(app_ctx, normalized_filters):
    now_ts = app_ctx.time.time()
    window = admin_metrics.resolve_period_window(normalized_filters.get('period', 'monthly'), now_ts=now_ts, runtime=app_ctx)
    pricing = admin_metrics.get_model_pricing_config(runtime=app_ctx)

    docs = admin_metrics.safe_query_docs_in_window(
        collection_name='job_logs',
        timestamp_field='finished_at',
        window_start=window['start'],
        window_end=window['end'],
        order_desc=True,
        filters=admin_metrics.admin_job_filters(runtime=app_ctx),
        runtime=app_ctx,
    )
    filtered_jobs = []
    for doc in docs:
        job = doc.to_dict() or {}
        job_id = str(job.get('job_id', doc.id) or doc.id)
        if not job_id:
            continue
        job['job_id'] = job_id
        if not admin_metrics.is_admin_visible_job(job, runtime=app_ctx):
            continue
        if _job_matches_filters(job, normalized_filters):
            filtered_jobs.append(job)

    selected_jobs = _select_jobs(filtered_jobs, normalized_filters)
    usd_to_eur = _to_non_negative_float(normalized_filters.get('usd_to_eur', 0.93), default=0.93) or 0.93

    job_rows = []
    stage_rows = []
    sum_input_tokens = 0
    sum_output_tokens = 0
    sum_total_tokens = 0
    sum_cost_usd = 0.0

    for job in selected_jobs:
        cost_info = admin_metrics.compute_job_stage_costs(job, pricing, runtime=app_ctx)
        job_input = _to_non_negative_int(cost_info.get('input_tokens', 0))
        job_output = _to_non_negative_int(cost_info.get('output_tokens', 0))
        job_total = _to_non_negative_int(cost_info.get('total_tokens', 0))
        job_cost_usd = float(cost_info.get('cost_usd', 0.0) or 0.0)
        job_cost_eur = job_cost_usd * usd_to_eur
        job_row = {
            'job_id': str(job.get('job_id', '') or ''),
            'uid': str(job.get('uid', '') or ''),
            'email': str(job.get('email', '') or ''),
            'mode': str(job.get('mode', '') or ''),
            'status': str(job.get('status', '') or ''),
            'finished_at': job.get('finished_at', 0),
            'billing_mode': str(job.get('billing_mode', 'standard') or 'standard'),
            'is_batch': bool(job.get('is_batch', False)),
            'batch_parent_id': str(job.get('batch_parent_id', '') or ''),
            'batch_row_id': str(job.get('batch_row_id', '') or ''),
            'token_input_total': job_input,
            'token_output_total': job_output,
            'token_total': job_total,
            'cost_usd': round(job_cost_usd, 8),
            'cost_eur': round(job_cost_eur, 8),
            'missing_stage_usage': bool(cost_info.get('missing_stage_usage', False)),
        }
        job_rows.append(job_row)

        for stage in cost_info.get('stages', []) or []:
            stage_rows.append(
                {
                    'job_id': job_row['job_id'],
                    'stage': str(stage.get('stage', '') or ''),
                    'model': str(stage.get('model', '') or ''),
                    'tier': str(stage.get('tier', '') or ''),
                    'billing_mode': str(stage.get('billing_mode', '') or ''),
                    'input_modality': str(stage.get('input_modality', '') or ''),
                    'input_tokens': _to_non_negative_int(stage.get('input_tokens', 0)),
                    'output_tokens': _to_non_negative_int(stage.get('output_tokens', 0)),
                    'total_tokens': _to_non_negative_int(stage.get('total_tokens', 0)),
                    'input_rate_per_million': float(stage.get('input_rate_per_million', 0.0) or 0.0),
                    'output_rate_per_million': float(stage.get('output_rate_per_million', 0.0) or 0.0),
                    'cost_input_usd': float(stage.get('cost_input_usd', 0.0) or 0.0),
                    'cost_output_usd': float(stage.get('cost_output_usd', 0.0) or 0.0),
                    'cost_usd': float(stage.get('cost_usd', 0.0) or 0.0),
                    'cost_eur': float(stage.get('cost_usd', 0.0) or 0.0) * usd_to_eur,
                    'matched_pricing': bool(stage.get('matched_pricing', False)),
                }
            )

        sum_input_tokens += job_input
        sum_output_tokens += job_output
        sum_total_tokens += job_total
        sum_cost_usd += job_cost_usd

    return {
        'filters': {
            'period': window['period'],
            'window_start': window['start'],
            'window_end': window['end'],
            'mode': normalized_filters.get('mode', ''),
            'status': normalized_filters.get('status', ''),
            'uid': normalized_filters.get('uid', ''),
            'email': normalized_filters.get('email', ''),
            'selection': normalized_filters.get('selection', 'all'),
            'selected_job_ids': list(normalized_filters.get('job_ids', []) or []),
        },
        'pricing_version': str((pricing or {}).get('version', '') or ''),
        'usd_to_eur': usd_to_eur,
        'summary': {
            'jobs_filtered': len(filtered_jobs),
            'jobs_selected': len(selected_jobs),
            'token_input_total': sum_input_tokens,
            'token_output_total': sum_output_tokens,
            'token_total': sum_total_tokens,
            'cost_usd_total': round(sum_cost_usd, 8),
            'cost_eur_total': round(sum_cost_usd * usd_to_eur, 8),
        },
        'jobs': job_rows,
        'stages': stage_rows,
    }


def admin_overview(app_ctx, request):
    return admin_dashboard_service.admin_overview(app_ctx, request)


def admin_export(app_ctx, request):
    return admin_dashboard_service.admin_export(app_ctx, request)


def admin_prompts(app_ctx, request):
    return admin_dashboard_service.admin_prompts(app_ctx, request)


def admin_model_pricing(app_ctx, request):
    return admin_dashboard_service.admin_model_pricing(app_ctx, request)


def admin_cost_analysis(app_ctx, request):
    _decoded, error_response, status = _require_admin(app_ctx, request)
    if error_response is not None:
        return error_response, status

    payload = request.get_json(silent=True) or {}
    normalized_filters = _normalize_analysis_filters(payload, runtime=app_ctx)
    try:
        analysis = _build_cost_analysis_payload(app_ctx, normalized_filters)
    except Exception as error:
        app_ctx.logger.error(f"Error building admin cost analysis: {error}")
        return app_ctx.jsonify({'error': 'Could not build cost analysis'}), 500
    return app_ctx.jsonify(analysis)


def admin_cost_analysis_export(app_ctx, request):
    _decoded, error_response, status = _require_admin(app_ctx, request)
    if error_response is not None:
        return error_response, status
    if Workbook is None:
        return app_ctx.jsonify({'error': 'Excel export dependency is missing'}), 500

    payload = request.get_json(silent=True) or {}
    normalized_filters = _normalize_analysis_filters(payload, runtime=app_ctx)
    try:
        analysis = _build_cost_analysis_payload(app_ctx, normalized_filters)
    except Exception as error:
        app_ctx.logger.error(f"Error building admin cost analysis export: {error}")
        return app_ctx.jsonify({'error': 'Could not export cost analysis'}), 500

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'

    summary_rows = [
        ['Pricing version', analysis.get('pricing_version', '')],
        ['Period', analysis.get('filters', {}).get('period', '')],
        ['Window start (UTC)', datetime.fromtimestamp(analysis.get('filters', {}).get('window_start', 0), tz=timezone.utc).isoformat()],
        ['Window end (UTC)', datetime.fromtimestamp(analysis.get('filters', {}).get('window_end', 0), tz=timezone.utc).isoformat()],
        ['Mode filter', analysis.get('filters', {}).get('mode', '') or 'all'],
        ['Status filter', analysis.get('filters', {}).get('status', '') or 'all'],
        ['UID filter', analysis.get('filters', {}).get('uid', '') or 'all'],
        ['Email filter', analysis.get('filters', {}).get('email', '') or 'all'],
        ['Selection mode', analysis.get('filters', {}).get('selection', 'all')],
        ['USD -> EUR rate', float(analysis.get('usd_to_eur', 0.93) or 0.93)],
        ['Jobs filtered', int((analysis.get('summary') or {}).get('jobs_filtered', 0) or 0)],
        ['Jobs selected', int((analysis.get('summary') or {}).get('jobs_selected', 0) or 0)],
        ['Input tokens total', int((analysis.get('summary') or {}).get('token_input_total', 0) or 0)],
        ['Output tokens total', int((analysis.get('summary') or {}).get('token_output_total', 0) or 0)],
        ['Token total', int((analysis.get('summary') or {}).get('token_total', 0) or 0)],
        ['Cost total (USD)', float((analysis.get('summary') or {}).get('cost_usd_total', 0.0) or 0.0)],
        ['Cost total (EUR)', float((analysis.get('summary') or {}).get('cost_eur_total', 0.0) or 0.0)],
    ]
    for row in summary_rows:
        summary_sheet.append([sanitize_excel_cell(value) for value in row])

    jobs_sheet = workbook.create_sheet('Jobs')
    jobs_sheet.append(
        [
            'job_id',
            'uid',
            'email',
            'mode',
            'status',
            'finished_at',
            'billing_mode',
            'is_batch',
            'batch_parent_id',
            'batch_row_id',
            'token_input_total',
            'token_output_total',
            'token_total',
            'cost_usd',
            'cost_eur',
            'missing_stage_usage',
        ]
    )
    for job in analysis.get('jobs', []) or []:
        jobs_sheet.append(
            [
                sanitize_excel_cell(job.get('job_id', '')),
                sanitize_excel_cell(job.get('uid', '')),
                sanitize_excel_cell(job.get('email', '')),
                sanitize_excel_cell(job.get('mode', '')),
                sanitize_excel_cell(job.get('status', '')),
                sanitize_excel_cell(job.get('finished_at', 0)),
                sanitize_excel_cell(job.get('billing_mode', '')),
                sanitize_excel_cell(bool(job.get('is_batch', False))),
                sanitize_excel_cell(job.get('batch_parent_id', '')),
                sanitize_excel_cell(job.get('batch_row_id', '')),
                sanitize_excel_cell(int(job.get('token_input_total', 0) or 0)),
                sanitize_excel_cell(int(job.get('token_output_total', 0) or 0)),
                sanitize_excel_cell(int(job.get('token_total', 0) or 0)),
                sanitize_excel_cell(float(job.get('cost_usd', 0.0) or 0.0)),
                sanitize_excel_cell(float(job.get('cost_eur', 0.0) or 0.0)),
                sanitize_excel_cell(bool(job.get('missing_stage_usage', False))),
            ]
        )

    stages_sheet = workbook.create_sheet('Stage Breakdown')
    stages_sheet.append(
        [
            'job_id',
            'stage',
            'model',
            'tier',
            'billing_mode',
            'input_modality',
            'input_tokens',
            'output_tokens',
            'total_tokens',
            'input_rate_per_million_usd',
            'output_rate_per_million_usd',
            'cost_input_usd',
            'cost_output_usd',
            'cost_usd',
            'cost_eur',
            'matched_pricing',
        ]
    )
    for stage in analysis.get('stages', []) or []:
        stages_sheet.append(
            [
                sanitize_excel_cell(stage.get('job_id', '')),
                sanitize_excel_cell(stage.get('stage', '')),
                sanitize_excel_cell(stage.get('model', '')),
                sanitize_excel_cell(stage.get('tier', '')),
                sanitize_excel_cell(stage.get('billing_mode', '')),
                sanitize_excel_cell(stage.get('input_modality', '')),
                sanitize_excel_cell(int(stage.get('input_tokens', 0) or 0)),
                sanitize_excel_cell(int(stage.get('output_tokens', 0) or 0)),
                sanitize_excel_cell(int(stage.get('total_tokens', 0) or 0)),
                sanitize_excel_cell(float(stage.get('input_rate_per_million', 0.0) or 0.0)),
                sanitize_excel_cell(float(stage.get('output_rate_per_million', 0.0) or 0.0)),
                sanitize_excel_cell(float(stage.get('cost_input_usd', 0.0) or 0.0)),
                sanitize_excel_cell(float(stage.get('cost_output_usd', 0.0) or 0.0)),
                sanitize_excel_cell(float(stage.get('cost_usd', 0.0) or 0.0)),
                sanitize_excel_cell(float(stage.get('cost_eur', 0.0) or 0.0)),
                sanitize_excel_cell(bool(stage.get('matched_pricing', False))),
            ]
        )

    output = app_ctx.io.BytesIO()
    workbook.save(output)
    output.seek(0)
    period = analysis.get('filters', {}).get('period', 'monthly')
    filename = f'admin-cost-analysis-{period}.xlsx'
    return app_ctx.send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def admin_user_search(app_ctx, request):
    _decoded, error_response, status = _require_admin(app_ctx, request)
    if error_response is not None:
        return error_response, status

    email = _normalize_admin_email(request.args.get('email', ''))
    if not email or '@' not in email:
        return app_ctx.jsonify({'error': 'Enter a valid email address.'}), 400

    try:
        user_doc, user_data = _find_user_doc_by_email(app_ctx, email)
    except Exception as error:
        app_ctx.logger.error('Admin user search failed for %s: %s', email, error)
        return app_ctx.jsonify({'error': 'Could not search users right now.'}), 500

    if user_doc is None:
        return app_ctx.jsonify({'error': 'No existing user was found for that email address.'}), 404

    uid = str(user_data.get('uid', '') or getattr(user_doc, 'id', '') or '').strip()
    if not uid:
        uid = str(getattr(user_doc, 'id', '') or '').strip()

    return app_ctx.jsonify({
        'user': _serialize_admin_credit_user(app_ctx, uid, user_data),
        'recent_grants': _list_admin_credit_grants(app_ctx, email=email, limit=20),
    })


def admin_credit_grants(app_ctx, request):
    _decoded, error_response, status = _require_admin(app_ctx, request)
    if error_response is not None:
        return error_response, status
    email = _normalize_admin_email(request.args.get('email', ''))
    uid = str(request.args.get('uid', '') or '').strip()
    limit = _to_non_negative_int(request.args.get('limit', 20), default=20)
    return app_ctx.jsonify({'grants': _list_admin_credit_grants(app_ctx, email=email, uid=uid, limit=limit)})


def admin_grant_user_credits(app_ctx, request, uid):
    decoded, error_response, status = _require_admin(app_ctx, request)
    if error_response is not None:
        return error_response, status
    safe_uid = str(uid or '').strip()
    if not safe_uid:
        return app_ctx.jsonify({'error': 'Missing user id.'}), 400

    payload = request.get_json(silent=True) or {}
    try:
        category_amounts = _coerce_admin_credit_amounts(payload)
    except ValueError as error:
        return app_ctx.jsonify({'error': str(error)}), 400

    grant_id = str(app_ctx.uuid.uuid4())
    now_ts = _admin_credit_now(app_ctx)
    actor = _admin_actor(decoded)
    note = _note_from_payload(payload)
    user_ref = app_ctx.users_repo.doc_ref(app_ctx.db, safe_uid)
    grant_ref = app_ctx.admin_credit_grants_repo.doc_ref(app_ctx.db, grant_id)

    try:
        snapshot = user_ref.get()
        if not snapshot.exists:
            return app_ctx.jsonify({'error': 'No existing user was found for that id.'}), 404
        user = snapshot.to_dict() or {}
        user.setdefault('uid', safe_uid)
        if _target_is_deleting(user):
            return app_ctx.jsonify({'error': 'This account is being deleted, so credits cannot be changed.'}), 409

        internal_credits = {
            ADMIN_CREDIT_FIELDS_BY_CATEGORY[category]: amount
            for category, amount in category_amounts.items()
        }
        updates = {
            field: app_ctx.firestore.Increment(amount)
            for field, amount in internal_credits.items()
        }
        updates['updated_at'] = now_ts

        email = str(user.get('email', '') or '').strip()
        email_normalized = _normalize_admin_email(user.get('email_normalized') or email)
        if email_normalized and user.get('email_normalized') != email_normalized:
            updates['email_normalized'] = email_normalized
            user['email_normalized'] = email_normalized

        for field, amount in internal_credits.items():
            user[field] = int(user.get(field, 0) or 0) + amount
        user['updated_at'] = now_ts

        record = {
            'grant_id': grant_id,
            'action': 'grant_credits',
            'uid': safe_uid,
            'email': email,
            'email_normalized': email_normalized,
            **actor,
            'credits': internal_credits,
            'credit_categories': dict(category_amounts),
            'unlimited_before': billing_credits.normalize_unlimited_credits(user.get('unlimited_credits')),
            'unlimited_after': billing_credits.normalize_unlimited_credits(user.get('unlimited_credits')),
            'note': note,
            'price_cents': 0,
            'currency': 'eur',
            'source': 'admin',
            'created_at': now_ts,
        }
        _commit_admin_credit_batch(app_ctx, user_ref, updates, grant_ref, record)
        updated_user = user
    except Exception as error:
        app_ctx.logger.error('Admin credit grant failed for %s: %s', safe_uid, error)
        return app_ctx.jsonify({'error': 'Could not grant credits right now.'}), 500

    return app_ctx.jsonify({
        'ok': True,
        'user': _serialize_admin_credit_user(app_ctx, safe_uid, updated_user),
        'grant': _serialize_admin_credit_record(grant_id, record),
    })


def admin_update_user_unlimited(app_ctx, request, uid):
    decoded, error_response, status = _require_admin(app_ctx, request)
    if error_response is not None:
        return error_response, status
    safe_uid = str(uid or '').strip()
    if not safe_uid:
        return app_ctx.jsonify({'error': 'Missing user id.'}), 400

    payload = request.get_json(silent=True) or {}
    try:
        unlimited_updates = _coerce_admin_unlimited_updates(payload)
    except ValueError as error:
        return app_ctx.jsonify({'error': str(error)}), 400

    grant_id = str(app_ctx.uuid.uuid4())
    now_ts = _admin_credit_now(app_ctx)
    actor = _admin_actor(decoded)
    note = _note_from_payload(payload)
    user_ref = app_ctx.users_repo.doc_ref(app_ctx.db, safe_uid)
    grant_ref = app_ctx.admin_credit_grants_repo.doc_ref(app_ctx.db, grant_id)

    try:
        snapshot = user_ref.get()
        if not snapshot.exists:
            return app_ctx.jsonify({'error': 'No existing user was found for that id.'}), 404
        user = snapshot.to_dict() or {}
        user.setdefault('uid', safe_uid)
        if _target_is_deleting(user):
            return app_ctx.jsonify({'error': 'This account is being deleted, so credits cannot be changed.'}), 409

        before = billing_credits.normalize_unlimited_credits(user.get('unlimited_credits'))
        after = dict(before)
        after.update(unlimited_updates)
        updates = {f'unlimited_credits.{category}': bool(value) for category, value in unlimited_updates.items()}
        updates['updated_at'] = now_ts

        email = str(user.get('email', '') or '').strip()
        email_normalized = _normalize_admin_email(user.get('email_normalized') or email)
        if email_normalized and user.get('email_normalized') != email_normalized:
            updates['email_normalized'] = email_normalized
            user['email_normalized'] = email_normalized

        user['unlimited_credits'] = after
        user['updated_at'] = now_ts
        record = {
            'grant_id': grant_id,
            'action': 'set_unlimited',
            'uid': safe_uid,
            'email': email,
            'email_normalized': email_normalized,
            **actor,
            'credits': {},
            'credit_categories': {},
            'unlimited_before': before,
            'unlimited_after': after,
            'note': note,
            'price_cents': 0,
            'currency': 'eur',
            'source': 'admin',
            'created_at': now_ts,
        }
        _commit_admin_credit_batch(app_ctx, user_ref, updates, grant_ref, record)
        updated_user = user
    except Exception as error:
        app_ctx.logger.error('Admin unlimited credit update failed for %s: %s', safe_uid, error)
        return app_ctx.jsonify({'error': 'Could not update unlimited access right now.'}), 500

    return app_ctx.jsonify({
        'ok': True,
        'user': _serialize_admin_credit_user(app_ctx, safe_uid, updated_user),
        'grant': _serialize_admin_credit_record(grant_id, record),
    })


def admin_batch_jobs(app_ctx, request):
    return admin_dashboard_service.admin_batch_jobs(app_ctx, request)
