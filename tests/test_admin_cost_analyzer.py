import io

import pytest
from openpyxl import load_workbook

from lecture_processor.domains.admin import metrics as admin_metrics
from tests.runtime_test_support import get_test_core

core = get_test_core()

pytestmark = pytest.mark.usefixtures('disable_sentry')


class _Doc:
    def __init__(self, doc_id, payload):
        self.id = doc_id
        self._payload = payload

    def to_dict(self):
        return dict(self._payload)


def _patch_admin_auth(monkeypatch):
    monkeypatch.setattr(core, 'verify_firebase_token', lambda _request: {'uid': 'admin-uid', 'email': 'admin@example.com'})
    monkeypatch.setattr(core, 'is_admin_user', lambda _decoded: True)


def _pricing_fixture():
    return {
        'version': 'test-pricing',
        'pricing_table': {
            'gemini-2.5-flash-lite': {
                'standard': {
                    'input_text_per_M': 0.10,
                    'input_audio_per_M': 0.30,
                    'output_per_M': 0.40,
                },
                'batch': {
                    'input_text_per_M': 0.05,
                    'input_audio_per_M': 0.15,
                    'output_per_M': 0.20,
                },
            },
            'gemini-2.5-pro': {
                'standard': {
                    'tiers': [
                        {'label': '<=200k', 'max_input_tokens': 200000, 'input_text_per_M': 1.25, 'input_audio_per_M': 1.25, 'output_per_M': 10.0},
                        {'label': '>200k', 'min_input_tokens': 200001, 'input_text_per_M': 2.5, 'input_audio_per_M': 2.5, 'output_per_M': 15.0},
                    ]
                },
                'batch': {
                    'tiers': [
                        {'label': '<=200k', 'max_input_tokens': 200000, 'input_text_per_M': 0.625, 'input_audio_per_M': 0.625, 'output_per_M': 5.0},
                        {'label': '>200k', 'min_input_tokens': 200001, 'input_text_per_M': 1.25, 'input_audio_per_M': 1.25, 'output_per_M': 7.5},
                    ]
                },
            },
        },
    }


def _job_docs_fixture(now_ts):
    return [
        _Doc(
            'job-1',
            {
                'job_id': 'job-1',
                'uid': 'u1',
                'email': 'u1@example.com',
                'mode': 'lecture-notes',
                'status': 'complete',
                'billing_mode': 'standard',
                'finished_at': now_ts,
                'token_usage_by_stage': {
                    'slide_extraction': {
                        'input_tokens': 100000,
                        'output_tokens': 20000,
                        'total_tokens': 120000,
                        'model': 'gemini-2.5-flash-lite',
                        'billing_mode': 'standard',
                        'input_modality': 'text',
                    },
                    'merge': {
                        'input_tokens': 50000,
                        'output_tokens': 10000,
                        'total_tokens': 60000,
                        'model': 'gemini-2.5-pro',
                        'billing_mode': 'standard',
                        'input_modality': 'text',
                    },
                },
            },
        ),
        _Doc(
            'job-2',
            {
                'job_id': 'job-2',
                'uid': 'u2',
                'email': 'u2@example.com',
                'mode': 'interview',
                'status': 'complete',
                'billing_mode': 'batch',
                'is_batch': True,
                'batch_parent_id': 'batch-abc',
                'batch_row_id': 'row-2',
                'finished_at': now_ts,
                'token_usage_by_stage': {
                    'interview_transcription': {
                        'input_tokens': 300000,
                        'output_tokens': 25000,
                        'total_tokens': 325000,
                        'model': 'gemini-2.5-pro',
                        'billing_mode': 'batch',
                        'input_modality': 'audio',
                    }
                },
            },
        ),
    ]


def test_admin_cost_analysis_contract(client, monkeypatch):
    _patch_admin_auth(monkeypatch)
    now_ts = core.time.time()
    monkeypatch.setattr(admin_metrics, 'get_model_pricing_config', lambda runtime=None: _pricing_fixture())
    monkeypatch.setattr(admin_metrics, 'safe_query_docs_in_window', lambda **_kwargs: _job_docs_fixture(now_ts))

    response = client.post(
        '/api/admin/cost-analysis',
        json={
            'period': 'monthly',
            'usd_to_eur': 0.92,
        },
    )

    assert response.status_code == 200
    body = response.get_json()
    assert body.get('pricing_version') == 'test-pricing'
    assert body.get('summary', {}).get('jobs_selected') == 2
    assert body.get('summary', {}).get('token_total', 0) > 0
    assert body.get('summary', {}).get('cost_usd_total', 0.0) > 0.0
    assert body.get('summary', {}).get('cost_eur_total', 0.0) > 0.0
    assert len(body.get('jobs', [])) == 2
    assert len(body.get('stages', [])) >= 2


def test_admin_cost_analysis_allows_job_selection(client, monkeypatch):
    _patch_admin_auth(monkeypatch)
    now_ts = core.time.time()
    monkeypatch.setattr(admin_metrics, 'get_model_pricing_config', lambda runtime=None: _pricing_fixture())
    monkeypatch.setattr(admin_metrics, 'safe_query_docs_in_window', lambda **_kwargs: _job_docs_fixture(now_ts))

    response = client.post(
        '/api/admin/cost-analysis',
        json={
            'period': 'monthly',
            'job_ids': ['job-1'],
            'usd_to_eur': 1.0,
        },
    )

    assert response.status_code == 200
    body = response.get_json()
    assert body.get('summary', {}).get('jobs_selected') == 1
    jobs = body.get('jobs', [])
    assert len(jobs) == 1
    assert jobs[0].get('job_id') == 'job-1'


def test_admin_cost_analysis_export_xlsx(client, monkeypatch):
    _patch_admin_auth(monkeypatch)
    now_ts = core.time.time()
    monkeypatch.setattr(admin_metrics, 'get_model_pricing_config', lambda runtime=None: _pricing_fixture())
    monkeypatch.setattr(admin_metrics, 'safe_query_docs_in_window', lambda **_kwargs: _job_docs_fixture(now_ts))

    response = client.post(
        '/api/admin/cost-analysis/export',
        json={
            'period': 'quarterly',
            'usd_to_eur': 0.9,
        },
    )

    assert response.status_code == 200
    assert response.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    workbook = load_workbook(io.BytesIO(response.data))
    assert workbook.sheetnames == ['Summary', 'Jobs', 'Stage Breakdown']
    jobs_sheet = workbook['Jobs']
    headers = [jobs_sheet.cell(row=1, column=idx).value for idx in range(1, 17)]
    assert 'token_input_total' in headers
    assert 'token_output_total' in headers
    assert 'token_total' in headers
    assert 'cost_usd' in headers
    assert 'cost_eur' in headers


def test_admin_jobs_export_includes_cost_columns(client, monkeypatch):
    _patch_admin_auth(monkeypatch)
    now_ts = core.time.time()
    monkeypatch.setattr(admin_metrics, 'get_model_pricing_config', lambda runtime=None: _pricing_fixture())
    monkeypatch.setattr(admin_metrics, 'safe_query_docs_in_window', lambda **_kwargs: _job_docs_fixture(now_ts))

    response = client.get('/api/admin/export?type=jobs&window=7d&usd_to_eur=1.0')

    assert response.status_code == 200
    csv_text = response.get_data(as_text=True)
    header = csv_text.splitlines()[0]
    assert 'token_input_total' in header
    assert 'token_output_total' in header
    assert 'token_total' in header
    assert 'cost_usd' in header
    assert 'cost_eur' in header
